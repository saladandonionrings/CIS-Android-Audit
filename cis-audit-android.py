#!/usr/bin/env python3
import argparse
import json
import subprocess
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def execute_adb_command(command):
    """Execute adb command and return its output or an error."""
    try:
        result = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        if result.returncode == 0:
            return result.stdout.strip(), None
        else:
            return None, result.stderr.strip()
    except subprocess.CalledProcessError as e:
        return None, str(e)


def execute_custom_check(device, setting):
    """Executes custom checks based on the setting name and determines compliance."""
    expected_value = setting["expected"]
    compliance, current_value = "Non-Compliant", None
    if expected_value == "Manual Verification":
        return "Requires Manual Verification", "To Verify", expected_value
    if setting['name'] == "get_max_users":
        command = f"adb -s {device} shell pm get-max-users"
    elif setting['name'] == "list_third_party_keyboards":
        command = f"adb -s {device} shell pm list packages -3 | grep 'keyboard'"
    elif setting['name'] == "encryption_status":
        command = f"adb -s {device} shell getprop ro.crypto.state"
    elif setting['name'] == "root_status":
        command = f"adb -s {device} shell su -c 'echo Rooted'"
    else:
        return "Unsupported custom check", "Non-Compliant", expected_value
    output, error = execute_adb_command(command)
    if not error:
        current_value = output
        if setting['name'] == "list_third_party_keyboards" and not output:
            compliance = "Compliant"
        elif output == expected_value:
            compliance = "Compliant"
        else:
            compliance = "Non-Compliant"
    else:
        compliance = "Error"
        current_value = error

    return current_value, compliance, expected_value

def compare_device_settings(device, setting_type, setting_name, expected_value):
    if expected_value == "Manual Verification":
        return "Requires Manual Verification", "To Verify"
    adb_command = f"adb -s {device} shell settings get {setting_type} {setting_name}"
    current_value, error = execute_adb_command(adb_command)
    if error:
        compliance = "Error"
        current_value = error
    else:
        compliance = "Compliant" if current_value == expected_value else "Non-Compliant"
    return current_value, compliance

# args
parser = argparse.ArgumentParser(description="Android Configuration Checker")
parser.add_argument('--device', required=True, help="Specify the device to audit")
parser.add_argument('--config', help="Path to the expected settings configuration file", default='settings.json')
args = parser.parse_args()

# settings.json
try:
    with open(args.config, 'r') as file:
        settings = json.load(file)
except FileNotFoundError:
    print(f"[x] Configuration file {args.config} not found.")
    exit(1)

wb = Workbook()
ws = wb.active
ws.title = "Settings Compliance"
headers = ["Description", "Category", "Setting Name", "Current Configuration", "Expected Configuration", "Compliance", "Steps"]
ws.append(headers)

# cell colors
fill_compliant = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
fill_non_compliant = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_error = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_manual_verification = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for setting_type in ['secure', 'global', 'custom_checks']:
    for setting in settings.get(setting_type, []):
        setting_name = setting["name"]
        description = setting["description"]
        expected_value = setting.get("expected", "Undefined")
        steps = setting.get("steps", "")
        if setting_type in ["secure", "global"]:
            current_value, compliance = compare_device_settings(args.device, setting_type, setting_name, expected_value)
        else:
            current_value, compliance, _ = execute_custom_check(args.device, setting)

        current_value = current_value if isinstance(current_value, str) else "Error or No Data"
        row = [description, setting_type, setting_name, current_value, expected_value, compliance, steps]
        ws.append(row)

        if compliance == "Compliant":
            fill = fill_compliant
        elif compliance == "Non-Compliant":
            fill = fill_non_compliant
        elif compliance == "To Verify":
            fill = fill_manual_verification
        else:
            fill = fill_error
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).fill = fill

# width columns
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

filename = f"Audit_Compliance_{args.device}.xlsx"
wb.save(filename)
print(f"Report saved to {filename}")
